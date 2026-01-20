#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Sanity test suite for your Ranked-Nodes Bayesian Network (Dom/Eco/Ling/AC -> AE).

What this script does
---------------------
1) Reads 13 scenarios from a CSV (default: scenarios_sanity_bn.csv).
2) For each scenario:
   - Sets evidence Dom, Eco, Ling, AC
   - Computes the posterior distribution of AE (VL..VH)
   - Computes AE_mean via centroids {0.1,0.3,0.5,0.7,0.9}
3) Saves:
   - out/sanity_bn_results.csv  (scenario-by-scenario outputs)
   - out/experimento_sanidade_bn.xlsx  (tabs: Cenarios, Resultados, Checks)
   - out/sanity_bn_checks.txt  (pass/fail checks + a compact summary)

Assumptions (match your current code snippets)
----------------------------------------------
- Your project has: Pipeline/evaluate_teams.py
- That module can provide a BN object, and the BN object supports:
    bn.setEvidence(<NodeName>, <StateLabel>)
    bn.calculateTPN(<NodeName>)   # returns distribution for that node
  (This is consistent with your current usage: bn.setEvidence(...) + bn.calculateTPN(...)).

If your BN API differs, edit ONLY the adapter functions:
  _get_bn(), _clear_evidence(), _set_evidence(), _calc_tpn()

Run
---
From the repository root (STFP):
  python3 experimento_sanidade_bn/run_sanity_bn.py

Optional:
  python3 experimento_sanidade_bn/run_sanity_bn.py \
    --input experimento_sanidade_bn/scenarios_sanity_bn.csv \
    --outdir experimento_sanidade_bn/out \
    --tol_mean 0.08
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

# -----------------------------
# Constants you can keep stable
# -----------------------------
STATES_ORDER: List[str] = ["VL", "L", "M", "H", "VH"]
DEFAULT_CENTROIDS: List[float] = [0.1, 0.3, 0.5, 0.7, 0.9]


@dataclass
class Scenario:
    scenario_id: str
    Dom: str
    Eco: str
    Ling: str
    AC: str
    group: str
    description: str


# =============================
# BN API ADAPTER (edit if needed)
# =============================

def _import_evaluate_teams(repo_root: Path):
    sys.path.insert(0, str(repo_root))
    try:
        from Pipeline import evaluate_teams as et  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "Could not import Pipeline.evaluate_teams. "
            "Run this from your repository root, or verify STFP/Pipeline/evaluate_teams.py exists. "
            f"Original error: {e}"
        )
    return et


def _get_bn(et: Any):
    """Try common access patterns to obtain the BN object."""
    for name in ("get_bn", "get_bn_cached", "get_bn_singleton", "get_bn_instance"):
        if hasattr(et, name) and callable(getattr(et, name)):
            return getattr(et, name)()

    # Sometimes people expose a singleton directly.
    for name in ("BN", "bn", "BN_SINGLETON"):
        if hasattr(et, name):
            obj = getattr(et, name)
            if obj is not None:
                return obj

    raise RuntimeError(
        "Could not obtain BN from Pipeline.evaluate_teams. "
        "Expose a function get_bn() (recommended), or set a module variable BN."
    )


def _clear_evidence(bn: Any) -> bool:
    """Attempt to clear evidence on the BN. Return True if it worked."""
    for name in ("clearEvidence", "clear_evidence", "clearEvidences", "resetEvidence", "reset_evidence"):
        if hasattr(bn, name) and callable(getattr(bn, name)):
            getattr(bn, name)()
            return True
    return False


def _set_evidence(bn: Any, node: str, state: str) -> None:
    """Set evidence for a node."""
    if hasattr(bn, "setEvidence") and callable(getattr(bn, "setEvidence")):
        bn.setEvidence(node, state)
        return

    # Alternative naming
    if hasattr(bn, "set_evidence") and callable(getattr(bn, "set_evidence")):
        try:
            bn.set_evidence({node: state})
        except TypeError:
            bn.set_evidence(node, state)
        return

    raise RuntimeError("BN object does not expose setEvidence(...) or set_evidence(...).")


def _calc_tpn(bn: Any, node: str) -> Sequence[float] | Dict[str, float] | Sequence[Tuple[str, float]]:
    """Return the posterior distribution for a node."""
    if hasattr(bn, "calculateTPN") and callable(getattr(bn, "calculateTPN")):
        return bn.calculateTPN(node)

    if hasattr(bn, "calculate_tpn") and callable(getattr(bn, "calculate_tpn")):
        return bn.calculate_tpn(node)

    raise RuntimeError("BN object does not expose calculateTPN(...) or calculate_tpn(...).")


# =============================
# Core logic
# =============================

def read_scenarios(csv_path: Path) -> List[Scenario]:
    scenarios: List[Scenario] = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        required = {"scenario_id", "Dom", "Eco", "Ling", "AC", "group", "description"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise ValueError(
                f"CSV missing columns. Required={sorted(required)}; got={reader.fieldnames}"
            )
        for row in reader:
            scenarios.append(
                Scenario(
                    scenario_id=row["scenario_id"].strip(),
                    Dom=row["Dom"].strip(),
                    Eco=row["Eco"].strip(),
                    Ling=row["Ling"].strip(),
                    AC=row["AC"].strip(),
                    group=row["group"].strip(),
                    description=row["description"].strip(),
                )
            )
    return scenarios


def normalize_dist(dist: Any) -> List[float]:
    """Normalize different distribution formats into a list aligned with STATES_ORDER."""
    # dict: {"VL":0.2,...}
    if isinstance(dist, dict):
        return [float(dist.get(k, 0.0)) for k in STATES_ORDER]

    # list of tuples: [("VL",0.2),...]
    if isinstance(dist, (list, tuple)) and dist and isinstance(dist[0], (list, tuple)) and len(dist[0]) == 2:
        d = {str(k): float(v) for k, v in dist}
        return [float(d.get(k, 0.0)) for k in STATES_ORDER]

    # plain list/tuple
    if isinstance(dist, (list, tuple)) and len(dist) == len(STATES_ORDER):
        return [float(x) for x in dist]

    raise ValueError(f"Unsupported distribution format: {type(dist)} -> {dist}")


def ae_mean_from_probs(probs: Sequence[float], centroids: Sequence[float]) -> float:
    return float(sum(p * c for p, c in zip(probs, centroids)))


def argmax_label(probs: Sequence[float]) -> Tuple[str, float]:
    idx = max(range(len(probs)), key=lambda i: probs[i])
    return STATES_ORDER[idx], float(probs[idx])


def run_suite(et: Any, scenarios: List[Scenario], centroids: List[float]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Run all scenarios. Returns (rows_for_csv, checks_report_lines)."""

    # Build BN once; if we can't clear evidence reliably, we'll rebuild per scenario.
    bn = _get_bn(et)
    can_clear = _clear_evidence(bn)

    results: List[Dict[str, Any]] = []

    for sc in scenarios:
        if not can_clear:
            bn = _get_bn(et)
        else:
            _clear_evidence(bn)

        _set_evidence(bn, "Dom", sc.Dom)
        _set_evidence(bn, "Eco", sc.Eco)
        _set_evidence(bn, "Ling", sc.Ling)
        _set_evidence(bn, "AC", sc.AC)

        dist_raw = _calc_tpn(bn, "AE")
        probs = normalize_dist(dist_raw)
        s = sum(probs)
        if s > 0:
            probs = [p / s for p in probs]

        mean = ae_mean_from_probs(probs, centroids)
        top_label, top_prob = argmax_label(probs)

        row: Dict[str, Any] = {
            "scenario_id": sc.scenario_id,
            "group": sc.group,
            "Dom": sc.Dom,
            "Eco": sc.Eco,
            "Ling": sc.Ling,
            "AC": sc.AC,
            "AE_p_VL": probs[0],
            "AE_p_L": probs[1],
            "AE_p_M": probs[2],
            "AE_p_H": probs[3],
            "AE_p_VH": probs[4],
            "AE_top": top_label,
            "AE_top_prob": top_prob,
            "AE_mean": mean,
            "description": sc.description,
        }
        results.append(row)

    # -------------------------
    # Checks / acceptance tests
    # -------------------------
    by_id = {r["scenario_id"]: r for r in results}

    def m(sid: str) -> float:
        return float(by_id[sid]["AE_mean"])

    def top(sid: str) -> str:
        return str(by_id[sid]["AE_top"])

    lines: List[str] = []
    lines.append("SANITY SUITE CHECKS")
    lines.append("===================")

    # Boundary checks: label-based (robust to non-one-hot)
    lines.append("\n[A] Boundary (extremes)")
    lines.append(f"S1 all VL -> AE_top={top('S1')} (expected VL), AE_mean={m('S1'):.4f}")
    lines.append(f"S2 all VH -> AE_top={top('S2')} (expected VH), AE_mean={m('S2'):.4f}")

    # Local monotonicity using VL <= M <= VH for each variable (3-point monotonic)
    lines.append("\n[B] Local monotonicity (3-point: VL <= M <= VH with others fixed at M)")

    checks = [
        ("Dom", "S3", "S0", "S4"),
        ("Eco", "S5", "S0", "S6"),
        ("Ling", "S7", "S0", "S8"),
        ("AC", "S9", "S0", "S10"),
    ]

    for var, sid_vl, sid_m, sid_vh in checks:
        ok = (m(sid_vl) <= m(sid_m) <= m(sid_vh))
        lines.append(
            f"{var}: mean({sid_vl})={m(sid_vl):.4f} <= mean({sid_m})={m(sid_m):.4f} <= mean({sid_vh})={m(sid_vh):.4f} -> {'PASS' if ok else 'FAIL'}"
        )

    # Trade-offs: should not saturate to full VH compared to all-VH case
    lines.append("\n[C] Trade-off (should not saturate to the all-VH ceiling)")
    lines.append(f"S11 tech=VH,VH,VH + AC=VL -> AE_mean={m('S11'):.4f} (must be < S2={m('S2'):.4f})")
    lines.append(f"S12 tech=VL,VL,VL + AC=VH -> AE_mean={m('S12'):.4f} (must be < S2={m('S2'):.4f})")

    lines.append("")
    return results, lines


def write_results_csv(rows: List[Dict[str, Any]], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_checks_txt(lines: List[str], out_txt: Path) -> None:
    out_txt.parent.mkdir(parents=True, exist_ok=True)
    out_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(scenarios: List[Scenario], rows: List[Dict[str, Any]], checks_lines: List[str], out_xlsx: Path) -> None:
    """Create an Excel file with 3 tabs: Cenarios, Resultados, Checks."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to write the .xlsx. Install it or remove the xlsx export. "
            f"Original error: {e}"
        )

    wb = Workbook()

    # ---- Sheet 1: scenarios ----
    ws1 = wb.active
    ws1.title = "Cenarios"
    headers1 = ["scenario_id", "Dom", "Eco", "Ling", "AC", "group", "description"]
    ws1.append(headers1)
    for sc in scenarios:
        ws1.append([sc.scenario_id, sc.Dom, sc.Eco, sc.Ling, sc.AC, sc.group, sc.description])

    # ---- Sheet 2: results ----
    ws2 = wb.create_sheet("Resultados")
    headers2 = list(rows[0].keys()) if rows else []
    ws2.append(headers2)
    for r in rows:
        ws2.append([r.get(h, "") for h in headers2])

    # ---- Sheet 3: checks ----
    ws3 = wb.create_sheet("Checks")
    ws3.append(["checks_report"])
    for line in checks_lines:
        ws3.append([line])

    # Simple formatting
    bold = Font(bold=True)
    for ws in (ws1, ws2, ws3):
        ws.freeze_panes = "A2"
        ws["A1"].font = bold
        # auto-width-ish
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = min(60, max(12, ws.column_dimensions[letter].width or 12))

    for cell in ws1[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    for cell in ws2[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    ws3["A1"].font = bold

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run BN sanity suite scenarios (13 cases).")
    parser.add_argument(
        "--input",
        type=str,
        default=str(Path(__file__).with_name("scenarios_sanity_bn.csv")),
        help="Input scenarios CSV (semicolon separated).",
    )
    parser.add_argument(
        "--outdir",
        type=str,
        default=str(Path(__file__).with_name("out")),
        help="Output directory.",
    )
    parser.add_argument(
        "--tol_mean",
        type=float,
        default=0.08,
        help="(Not used for hard FAIL) Suggested tolerance for 'near centroid' language in the thesis.",
    )

    args = parser.parse_args()

    input_csv = Path(args.input).resolve()
    outdir = Path(args.outdir).resolve()

    # Repo root is one level above this folder (STFP/experimento_sanidade_bn)
    repo_root = Path(__file__).resolve().parents[1]

    et = _import_evaluate_teams(repo_root)

    centroids = DEFAULT_CENTROIDS
    if hasattr(et, "CENTROIDES"):
        try:
            c = list(getattr(et, "CENTROIDES"))
            if len(c) == 5:
                centroids = [float(x) for x in c]
        except Exception:
            pass

    scenarios = read_scenarios(input_csv)
    rows, checks_lines = run_suite(et, scenarios, centroids)

    out_csv = outdir / "sanity_bn_results.csv"
    out_xlsx = outdir / "experimento_sanidade_bn.xlsx"
    out_txt = outdir / "sanity_bn_checks.txt"

    write_results_csv(rows, out_csv)
    write_checks_txt(checks_lines, out_txt)
    write_xlsx(scenarios, rows, checks_lines, out_xlsx)

    # Console summary
    print("\n".join(checks_lines))
    print("\nOutputs written:")
    print(f"- {out_csv}")
    print(f"- {out_xlsx}")
    print(f"- {out_txt}")

    # Friendly thesis hint
    print("\nThesis wording hint (use tol_mean):")
    print(
        f"Use wording like: 'AE_mean close to the VL centroid (0.1) within ±{args.tol_mean:.2f}' "
        "instead of requiring an exact one-hot distribution."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
